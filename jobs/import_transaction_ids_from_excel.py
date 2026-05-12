"""
用途：解决新轮询系统中1000页后无法访问的数据，使用excel进行导入

从 Excel 第一行表头识别列：「A端订单号」（或「订单编号」）对应 orders.order_id，
「平台订单号」（或「交易号」）写入 orders.transaction_id。
用 Excel 中的订单编号在库里匹配：order_id 包含该片段（LOCATE，等价于 LIKE '%片段%'）；若命中多个不同的 order_id 则跳过；若多条记录实为同一 order_id 则正常更新。
匹配不上的行跳过，不插入新订单。

依赖：pip install openpyxl（见 jobs/requirements.txt）

用法：
  python3 jobs/import_transaction_ids_from_excel.py --excel /path/to/file.xlsx
  python3 jobs/import_transaction_ids_from_excel.py --excel ./data.xlsx --dry-run
"""
from __future__ import annotations

import argparse
import logging
import os
from typing import Any, Dict, Tuple

import pymysql
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

MYSQL_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "port": int(os.getenv("DB_PORT", 3306)),
    "user": os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", ""),
    "database": os.getenv("DB_NAME", ""),
    "charset": "utf8mb4",
}


def normalize_header(text: Any) -> str:
    """标准化表头，用于与候选列名比较"""
    if text is None:
        return ""
    return str(text).strip().lower().replace(" ", "").replace("_", "")


def ensure_transaction_id_column(cursor) -> None:
    db_name = MYSQL_CONFIG["database"]
    cursor.execute(
        """
        SELECT COUNT(1) FROM information_schema.columns
        WHERE table_schema = %s AND table_name = 'orders' AND column_name = 'transaction_id'
        """,
        (db_name,),
    )
    if cursor.fetchone()[0] == 0:
        cursor.execute(
            "ALTER TABLE orders ADD COLUMN transaction_id VARCHAR(100) DEFAULT NULL "
            "COMMENT '交易号/平台订单号' AFTER order_id"
        )
        logger.info("已为 orders 表新增 transaction_id 字段")


def resolve_columns(header_row: Tuple[Any, ...]) -> Tuple[int, int]:
    """
    根据实际表头返回 (order_col_index, txn_col_index)
    优先精确匹配原始字符串，再尝试标准化匹配。
    """
    raw_headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    norm_headers = [normalize_header(cell) for cell in raw_headers]

    order_candidates_raw = ["A端订单号", "订单编号"]
    txn_candidates_raw = ["平台订单号", "交易号"]

    order_candidates_norm = [normalize_header(c) for c in order_candidates_raw]
    txn_candidates_norm = [normalize_header(c) for c in txn_candidates_raw]

    order_col = None
    txn_col = None

    for idx, raw in enumerate(raw_headers):
        if raw in order_candidates_raw:
            order_col = idx
            break
    for idx, raw in enumerate(raw_headers):
        if raw in txn_candidates_raw:
            txn_col = idx
            break

    if order_col is None:
        for idx, norm in enumerate(norm_headers):
            if norm in order_candidates_norm:
                order_col = idx
                break
    if txn_col is None:
        for idx, norm in enumerate(norm_headers):
            if norm in txn_candidates_norm:
                txn_col = idx
                break

    if order_col is None:
        raise ValueError(f"表头中未找到「A端订单号」或「订单编号」列，实际表头: {raw_headers}")
    if txn_col is None:
        raise ValueError(f"表头中未找到「平台订单号」或「交易号」列，实际表头: {raw_headers}")

    return order_col, txn_col


def import_excel(excel_path: str, dry_run: bool) -> Dict[str, int]:
    if not os.path.isfile(excel_path):
        raise FileNotFoundError(f"文件不存在: {excel_path}")

    # 不使用 read_only 模式，避免合并单元格读取异常
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    # 直接获取第一行所有单元格的值
    first_row = ws[1]
    header_row = [cell.value for cell in first_row]
    if not any(header_row):
        wb.close()
        raise ValueError("Excel 第一行为空")

    order_col, txn_col = resolve_columns(tuple(header_row))

    pairs: Dict[str, str] = {}
    data_rows = 0
    skipped_empty = 0

    # 从第二行开始读取数据
    for row in ws.iter_rows(min_row=2, values_only=True):
        data_rows += 1
        if not row:
            skipped_empty += 1
            continue
        oid = (
            str(row[order_col]).strip()
            if order_col < len(row) and row[order_col] is not None
            else ""
        )
        tid = (
            str(row[txn_col]).strip()
            if txn_col < len(row) and row[txn_col] is not None
            else ""
        )
        if not oid:
            skipped_empty += 1
            continue
        if tid:
            pairs[oid] = tid

    wb.close()

    stats = {
        "excel_data_rows": data_rows,
        "skipped_empty_or_no_order": skipped_empty,
        "unique_order_ids_with_txn": len(pairs),
        "db_updated": 0,
        "db_no_match": 0,
        "db_ambiguous": 0,
    }

    if dry_run:
        logger.info("[dry-run] 将更新 %s 个 order_id 的 transaction_id（去重后）", len(pairs))
        return stats

    conn = pymysql.connect(**MYSQL_CONFIG)
    cursor = conn.cursor()
    try:
        ensure_transaction_id_column(cursor)
        conn.commit()

        # 与 order_id LIKE '%Excel片段%' 等价；DISTINCT 避免同一 order_id 多行时误判为「多条」
        select_sql = (
            "SELECT DISTINCT order_id FROM orders WHERE LOCATE(%s, order_id) > 0 LIMIT 3"
        )
        update_sql = "UPDATE orders SET transaction_id = %s WHERE order_id = %s"
        for order_id, txn in pairs.items():
            tid = txn[:100] if len(txn) > 100 else txn
            if len(txn) > 100:
                logger.warning("平台订单号超长已截断：excel_order=%s len=%s", order_id, len(txn))
            cursor.execute(select_sql, (order_id,))
            hits = [row[0] for row in cursor.fetchall()]
            unique_ids = list(dict.fromkeys(hits))
            if len(unique_ids) == 0:
                stats["db_no_match"] += 1
            elif len(unique_ids) == 1:
                cursor.execute(update_sql, (tid, unique_ids[0]))
                if cursor.rowcount > 0:
                    stats["db_updated"] += cursor.rowcount
            else:
                stats["db_ambiguous"] += 1
                logger.warning(
                    "订单号模糊匹配到多个不同 order_id，已跳过：excel=%s matches=%s",
                    order_id,
                    unique_ids[:5],
                )

        conn.commit()
    finally:
        cursor.close()
        conn.close()

    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel 平台订单号 -> orders.transaction_id")
    parser.add_argument("--excel", "-e", required=True, help="Excel 文件路径 (.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="只解析 Excel，不写数据库")
    args = parser.parse_args()

    if not MYSQL_CONFIG["database"] and not args.dry_run:
        logger.error("请配置 DB_NAME")
        raise SystemExit(1)

    stats = import_excel(args.excel, dry_run=args.dry_run)
    logger.info(
        "完成：excel_data_rows=%s skipped=%s unique_pairs=%s db_updated=%s db_no_match=%s db_ambiguous=%s dry_run=%s",
        stats["excel_data_rows"],
        stats["skipped_empty_or_no_order"],
        stats["unique_order_ids_with_txn"],
        stats["db_updated"],
        stats["db_no_match"],
        stats.get("db_ambiguous", 0),
        args.dry_run,
    )


if __name__ == "__main__":
    main()